# -*- coding: utf-8 -*-
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import json
import base64


class AnalyticsReport(models.Model):
    _name = 'analytics.report'
    _description = 'Analytics Report'
    _order = 'create_date desc'

    name = fields.Char(string='Report Name', required=True)
    
    report_type = fields.Selection([
        ('standard', 'Standard Report'),
        ('custom', 'Custom Report'),
        ('scheduled', 'Scheduled Report'),
    ], string='Report Type', required=True, default='standard')
    
    category = fields.Selection([
        ('workforce', 'Workforce Report'),
        ('payroll', 'Payroll Report'),
        ('recruitment', 'Recruitment Report'),
        ('compliance', 'Compliance Report'),
        ('financial', 'Financial Report'),
        ('executive', 'Executive Summary'),
    ], string='Category', required=True)
    
    description = fields.Text(string='Description')
    
    # Date Range
    date_range = fields.Selection([
        ('today', 'Today'),
        ('yesterday', 'Yesterday'),
        ('this_week', 'This Week'),
        ('last_week', 'Last Week'),
        ('this_month', 'This Month'),
        ('last_month', 'Last Month'),
        ('this_quarter', 'This Quarter'),
        ('last_quarter', 'Last Quarter'),
        ('this_year', 'This Year'),
        ('last_year', 'Last Year'),
        ('custom', 'Custom Range'),
    ], string='Date Range', default='this_month')
    
    date_from = fields.Date(string='Date From')
    date_to = fields.Date(string='Date To')
    
    # Filters
    department_ids = fields.Many2many('hr.department', string='Departments')
    company_ids = fields.Many2many('res.company', string='Companies')
    
    # Report Configuration
    include_charts = fields.Boolean(string='Include Charts', default=True)
    include_tables = fields.Boolean(string='Include Tables', default=True)
    include_summary = fields.Boolean(string='Include Summary', default=True)
    
    # KPIs to include
    kpi_ids = fields.Many2many('analytics.kpi', string='KPIs')
    
    # Output
    output_format = fields.Selection([
        ('pdf', 'PDF'),
        ('excel', 'Excel'),
        ('html', 'HTML'),
    ], string='Output Format', default='pdf')
    
    # Generated Report
    report_file = fields.Binary(string='Report File')
    report_filename = fields.Char(string='Filename')
    
    # Scheduling
    is_scheduled = fields.Boolean(string='Scheduled')
    schedule_frequency = fields.Selection([
        ('daily', 'Daily'),
        ('weekly', 'Weekly'),
        ('monthly', 'Monthly'),
        ('quarterly', 'Quarterly'),
    ], string='Frequency')
    schedule_day = fields.Integer(string='Day of Month/Week', default=1)
    schedule_time = fields.Float(string='Time')
    recipient_ids = fields.Many2many('res.users', string='Recipients')
    next_run = fields.Datetime(string='Next Run')
    last_run = fields.Datetime(string='Last Run')
    
    company_id = fields.Many2one('res.company', string='Company', 
                                  default=lambda self: self.env.company)

    @api.onchange('date_range')
    def _onchange_date_range(self):
        """Set date range based on selection."""
        today = fields.Date.today()
        
        if self.date_range == 'today':
            self.date_from = self.date_to = today
        elif self.date_range == 'yesterday':
            self.date_from = self.date_to = today - timedelta(days=1)
        elif self.date_range == 'this_week':
            self.date_from = today - timedelta(days=today.weekday())
            self.date_to = today
        elif self.date_range == 'last_week':
            self.date_from = today - timedelta(days=today.weekday() + 7)
            self.date_to = today - timedelta(days=today.weekday() + 1)
        elif self.date_range == 'this_month':
            self.date_from = today.replace(day=1)
            self.date_to = today
        elif self.date_range == 'last_month':
            last_month = today - relativedelta(months=1)
            self.date_from = last_month.replace(day=1)
            self.date_to = today.replace(day=1) - timedelta(days=1)
        elif self.date_range == 'this_quarter':
            quarter_start_month = ((today.month - 1) // 3) * 3 + 1
            self.date_from = today.replace(month=quarter_start_month, day=1)
            self.date_to = today
        elif self.date_range == 'this_year':
            self.date_from = today.replace(month=1, day=1)
            self.date_to = today

    def action_generate_report(self):
        """Generate the report."""
        self.ensure_one()
        
        # Get report data
        data = self._get_report_data()
        
        if self.output_format == 'excel':
            self._generate_excel_report(data)
        elif self.output_format == 'pdf':
            self._generate_pdf_report(data)
        else:
            self._generate_html_report(data)
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'analytics.report',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def _get_report_data(self):
        """Gather all data for the report."""
        data = {
            'report': {
                'name': self.name,
                'category': self.category,
                'date_from': self.date_from,
                'date_to': self.date_to,
                'generated_at': fields.Datetime.now(),
                'generated_by': self.env.user.name,
            },
            'summary': {},
            'kpis': [],
            'tables': [],
            'charts': [],
        }
        
        # Get summary data
        if self.include_summary:
            data['summary'] = self._get_summary_data()
        
        # Get KPI data
        for kpi in self.kpi_ids:
            data['kpis'].append(kpi.get_kpi_data())
        
        # Get table data based on category
        if self.include_tables:
            data['tables'] = self._get_table_data()
        
        return data

    def _get_summary_data(self):
        """Get summary statistics for the report."""
        Employee = self.env['hr.employee'].sudo()
        
        return {
            'total_employees': Employee.search_count([('active', '=', True)]),
            'new_hires': Employee.search_count([
                ('create_date', '>=', self.date_from),
                ('create_date', '<=', self.date_to),
            ]),
            'departures': Employee.search_count([
                ('departure_date', '>=', self.date_from),
                ('departure_date', '<=', self.date_to),
            ]),
        }

    def _get_table_data(self):
        """Get table data for the report."""
        tables = []
        
        if self.category == 'workforce':
            tables.append(self._get_employee_table())
        elif self.category == 'payroll':
            tables.append(self._get_payroll_table())
        elif self.category == 'recruitment':
            tables.append(self._get_recruitment_table())
        
        return tables

    def _get_employee_table(self):
        """Get employee data table."""
        Employee = self.env['hr.employee'].sudo()
        employees = Employee.search([('active', '=', True)], limit=100)
        
        rows = []
        for emp in employees:
            rows.append({
                'name': emp.name,
                'department': emp.department_id.name or '-',
                'job': emp.job_id.name or '-',
                'hire_date': emp.create_date.strftime('%Y-%m-%d') if emp.create_date else '-',
            })
        
        return {
            'title': 'Employee List',
            'headers': ['Name', 'Department', 'Job Position', 'Hire Date'],
            'rows': rows,
        }

    def _get_payroll_table(self):
        """Get payroll data table."""
        return {
            'title': 'Payroll Summary',
            'headers': ['Department', 'Employees', 'Total Cost', 'Average'],
            'rows': [],
        }

    def _get_recruitment_table(self):
        """Get recruitment data table."""
        return {
            'title': 'Recruitment Summary',
            'headers': ['Position', 'Applications', 'Interviews', 'Offers', 'Hired'],
            'rows': [],
        }

    def _generate_excel_report(self, data):
        """Generate Excel report."""
        import io
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise UserError(_('openpyxl library is required for Excel export.'))
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Report'
        
        # Title
        ws['A1'] = data['report']['name']
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Date range
        ws['A2'] = f"Period: {data['report']['date_from']} to {data['report']['date_to']}"
        ws.merge_cells('A2:D2')
        
        row = 4
        
        # KPIs
        if data['kpis']:
            ws[f'A{row}'] = 'Key Performance Indicators'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for kpi in data['kpis']:
                ws[f'A{row}'] = kpi['name']
                ws[f'B{row}'] = kpi['display']
                row += 1
            
            row += 1
        
        # Tables
        for table in data.get('tables', []):
            ws[f'A{row}'] = table['title']
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            # Headers
            for col, header in enumerate(table['headers'], 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', fill_type='solid')
            row += 1
            
            # Data rows
            for data_row in table['rows']:
                col = 1
                for key in ['name', 'department', 'job', 'hire_date']:
                    if key in data_row:
                        ws.cell(row=row, column=col, value=data_row[key])
                        col += 1
                row += 1
            
            row += 1
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        self.report_file = base64.b64encode(buffer.read())
        self.report_filename = f"{self.name.replace(' ', '_')}_{fields.Date.today()}.xlsx"

    def _generate_pdf_report(self, data):
        """Generate PDF report."""
        # Would use reportlab or weasyprint
        # For now, create a simple text-based report
        content = f"""
TAZWEED HR ANALYTICS REPORT
===========================

Report: {data['report']['name']}
Category: {data['report']['category']}
Period: {data['report']['date_from']} to {data['report']['date_to']}
Generated: {data['report']['generated_at']}
Generated By: {data['report']['generated_by']}

KEY PERFORMANCE INDICATORS
--------------------------
"""
        for kpi in data['kpis']:
            content += f"{kpi['name']}: {kpi['display']}\n"
        
        content += "\n"
        
        for table in data.get('tables', []):
            content += f"\n{table['title']}\n"
            content += "-" * 50 + "\n"
            content += " | ".join(table['headers']) + "\n"
            content += "-" * 50 + "\n"
            for row in table['rows'][:10]:
                content += " | ".join(str(v) for v in row.values()) + "\n"
        
        self.report_file = base64.b64encode(content.encode('utf-8'))
        self.report_filename = f"{self.name.replace(' ', '_')}_{fields.Date.today()}.txt"

    def _generate_html_report(self, data):
        """Generate HTML report."""
        html = f"""
<!DOCTYPE html>
<html>
<head>
    <title>{data['report']['name']}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 40px; }}
        h1 {{ color: #2196F3; }}
        .kpi-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 20px; margin: 20px 0; }}
        .kpi-card {{ background: #f5f5f5; padding: 20px; border-radius: 8px; text-align: center; }}
        .kpi-value {{ font-size: 24px; font-weight: bold; color: #2196F3; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
        th {{ background: #2196F3; color: white; }}
    </style>
</head>
<body>
    <h1>{data['report']['name']}</h1>
    <p>Period: {data['report']['date_from']} to {data['report']['date_to']}</p>
    
    <h2>Key Performance Indicators</h2>
    <div class="kpi-grid">
"""
        for kpi in data['kpis']:
            html += f"""
        <div class="kpi-card">
            <div class="kpi-name">{kpi['name']}</div>
            <div class="kpi-value">{kpi['display']}</div>
        </div>
"""
        
        html += "</div>"
        
        for table in data.get('tables', []):
            html += f"<h2>{table['title']}</h2><table><tr>"
            for header in table['headers']:
                html += f"<th>{header}</th>"
            html += "</tr>"
            for row in table['rows'][:20]:
                html += "<tr>"
                for val in row.values():
                    html += f"<td>{val}</td>"
                html += "</tr>"
            html += "</table>"
        
        html += "</body></html>"
        
        self.report_file = base64.b64encode(html.encode('utf-8'))
        self.report_filename = f"{self.name.replace(' ', '_')}_{fields.Date.today()}.html"

    def action_download_report(self):
        """Download the generated report."""
        self.ensure_one()
        
        if not self.report_file:
            raise UserError(_('Please generate the report first.'))
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/analytics.report/{self.id}/report_file/{self.report_filename}?download=true',
            'target': 'new',
        }
